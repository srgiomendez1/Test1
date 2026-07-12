"""
Flagstar Treasury / ALM - weekly brokered-CD term premium engine.

Mirrors the Term Premium Monitor dashboard exactly:

    market_TP(t) = max( mean_panel[ CD_offer(t) ], FSB_issuance(t) ) - SOFR(t)

Governing comparison: production (approved TP schedule) vs the rolling
LOOKBACK-week trend of market_TP INCLUDING the current week. A tenor is
"Discuss" only when that trend leaves the +/- BAND bp zone of indifference;
a single-week breach alone is "Watch".

Pure-Python + openpyxl; runs identically on Windows, macOS and Linux.
"""
from __future__ import annotations

import csv
import datetime as dt
import re
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook

TERMS = [(1, "1M"), (3, "3M"), (6, "6M"), (12, "1Y"),
         (24, "2Y"), (36, "3Y"), (48, "4Y"), (60, "5Y")]
TERM_MONTHS = [m for m, _ in TERMS]
TERM_LABELS = [l for _, l in TERMS]
NT = len(TERMS)
MMAX = 60

# Committee-advised TP schedule (bp) - Rate Summary "FCS" column, week of 2026-07-06.
DEFAULT_APPROVED = [15, 20, 25, 30, 50, 55, 60, 60]
DEFAULT_BAND_BP = 20
DEFAULT_LOOKBACK = 13

# ---------------------------------------------------------------- parsing ---

def norm_term(value) -> int | None:
    """'3M' / '6 mo' / '1Y' / '2 yr' -> months; anything else -> None."""
    if value is None:
        return None
    s = str(value).strip().lower()
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*[- ]?\s*(m|mo|mos|mon|month|months)\.?", s)
    if m:
        return round(float(m.group(1)))
    m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*[- ]?\s*(y|yr|yrs|year|years)\.?", s)
    if m:
        return round(float(m.group(1)) * 12)
    return None


def norm_date(value) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)) and 20000 < value < 60000:   # Excel serial
        return (dt.datetime(1899, 12, 30) + dt.timedelta(days=float(value))).date()
    s = str(value).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)                # ISO
    if m:
        return dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})", s)              # M/D/YYYY
    if m:
        y = int(m.group(3))
        return dt.date(y + 2000 if y < 100 else y, int(m.group(1)), int(m.group(2)))
    return None


def norm_rate(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        s = value.replace("%", "").strip()
        if re.fullmatch(r"-?\d+,\d+", s):        # decimal-comma locales
            s = s.replace(",", ".")
        try:
            value = float(s)
        except ValueError:
            return None
    v = float(value)
    if 0 < v < 1:                                # percent-formatted cell (0.0428)
        v *= 100
    return round(v, 3)


def read_rows(path: Path) -> list[list]:
    """First worksheet of an .xlsx/.xlsm (prefers a sheet named 'Data'),
    or a delimited text file, as a list of rows."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    text = path.read_text(encoding="utf-8-sig")   # strips Windows BOM
    delim = "\t" if "\t" in text else (";" if ";" in text and "," not in text else ",")
    return [row for row in csv.reader(text.splitlines(), delimiter=delim) if any(c.strip() for c in row)]


def _find_header(rows):
    for ri, row in enumerate(rows[:6]):
        hdr = [str(c or "").strip() for c in row]
        has_date = any(re.fullmatch(r"(?i)date|as ?of|observation ?date|effective ?date|print ?date", h) for h in hdr)
        term_cols = sum(1 for h in hdr if norm_term(h) is not None)
        has_long = (any(re.fullmatch(r"(?i)term|tenor|maturity", h) for h in hdr)
                    and any(re.match(r"(?i)rate|yield|level|apy|coupon", h) for h in hdr))
        if has_date and (term_cols >= 2 or has_long):
            return ri
    raise ValueError("No header row with a Date column plus term columns (or Term/Rate) found.")


def _col(hdr, pattern):
    for i, x in enumerate(hdr):
        if re.fullmatch(pattern, str(x or "").strip()):
            return i
    return -1


def load_cd(path: Path):
    """Brokered CD panel: Date | Issuer | term columns. Latest date wins.
    Returns (date, [(issuer, [rate-or-None per term]), ...])."""
    rows = read_rows(path)
    hi = _find_header(rows)
    hdr = [str(c or "").strip() for c in rows[hi]]
    di = _col(hdr, r"(?i)date|as ?of|observation ?date|effective ?date")
    ii = _col(hdr, r"(?i)issuer|bank|name|institution")
    if ii < 0:
        raise ValueError("Brokered CD file needs an Issuer column.")
    tcols = [(ci, norm_term(hdr[ci])) for ci in range(len(hdr)) if norm_term(hdr[ci]) is not None]
    recs = []
    for row in rows[hi + 1:]:
        d = norm_date(row[di]) if di < len(row) else None
        name = str(row[ii] or "").strip() if ii < len(row) else ""
        if not d or not name:
            continue
        offers = [None] * NT
        for ci, months in tcols:
            if months in TERM_MONTHS and ci < len(row):
                offers[TERM_MONTHS.index(months)] = norm_rate(row[ci])
        recs.append((d, name, offers))
    if not recs:
        raise ValueError("No CD rows with a valid date and issuer.")
    latest = max(r[0] for r in recs)
    panel = [(name, offers) for d, name, offers in recs if d == latest]
    return latest, panel


def load_wide(path: Path):
    """SOFR / FHLB: Date | term columns. Latest value per term.
    Returns (date, [rate-or-None per term])."""
    rows = read_rows(path)
    hi = _find_header(rows)
    hdr = [str(c or "").strip() for c in rows[hi]]
    di = _col(hdr, r"(?i)date|as ?of|observation ?date|effective ?date")
    tcols = [(ci, norm_term(hdr[ci])) for ci in range(len(hdr)) if norm_term(hdr[ci]) is not None]
    out, used = [None] * NT, []
    dated = [r for r in rows[hi + 1:] if di < len(r) and norm_date(r[di]) is not None]
    for row in sorted(dated, key=lambda r: norm_date(r[di])):
        d = norm_date(row[di])
        for ci, months in tcols:
            if months in TERM_MONTHS and ci < len(row):
                v = norm_rate(row[ci])
                if v is not None:
                    out[TERM_MONTHS.index(months)] = v
                    used.append(d)
    if sum(v is not None for v in out) < 3:
        raise ValueError(f"{path.name}: fewer than 3 terms loaded - check headers.")
    return (max(used) if used else None), out


def load_fsb(path: Path | None):
    """FSB issuances: long (Date|Term|Rate) or wide. Latest print per term.
    Missing/blank file => no floor. Returns (date-or-None, [rate-or-None per term])."""
    empty = (None, [None] * NT)
    if path is None or not Path(path).exists():
        return empty
    rows = read_rows(path)
    try:
        hi = _find_header(rows)
    except ValueError:
        return empty
    hdr = [str(c or "").strip() for c in rows[hi]]
    di = _col(hdr, r"(?i)date|as ?of|print ?date|effective ?date")
    ti = _col(hdr, r"(?i)term|tenor|maturity")
    ri = -1
    for i, x in enumerate(hdr):
        if re.match(r"(?i)rate|yield|level|apy|coupon", str(x or "").strip()):
            ri = i
            break
    out, used = [None] * NT, []
    if ti >= 0 and ri >= 0:      # long format
        recs = []
        for row in rows[hi + 1:]:
            d = norm_date(row[di]) if di < len(row) else None
            months = norm_term(row[ti]) if ti < len(row) else None
            v = norm_rate(row[ri]) if ri < len(row) else None
            if d and months in TERM_MONTHS and v is not None:
                recs.append((d, months, v))
        for d, months, v in sorted(recs):
            out[TERM_MONTHS.index(months)] = v
            used.append(d)
        return (max(used) if used else None), out
    try:                          # wide format
        return load_wide(path)
    except ValueError:
        return empty


def load_history(path: Path):
    """history.csv: date,1M,3M,...,5Y (bp). Returns [(date, [bp...]), ...] oldest first."""
    path = Path(path)
    if not path.exists():
        return []
    out = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            d = norm_date(row[0]) if row else None
            if d is None:
                continue                                   # header / junk
            vals = [(float(x) if x.strip() else None) for x in row[1:NT + 1]]
            vals += [None] * (NT - len(vals))
            out.append((d, vals))
    return sorted(out, key=lambda r: r[0])


def save_history(path: Path, history):
    with Path(path).open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["date"] + TERM_LABELS)
        for d, vals in history:
            w.writerow([d.isoformat()] + ["" if v is None else v for v in vals])

# ---------------------------------------------------------------- compute ---

def _mean(vals):
    xs = [v for v in vals if v is not None]
    return mean(xs) if xs else None


def interp_monthly(node_vals):
    """Piecewise linear on the node grid, flat outside - matches the FTP engine."""
    nodes = [(m, v) for (m, _), v in zip(TERMS, node_vals) if v is not None]
    out = []
    for month in range(1, MMAX + 1):
        if not nodes:
            out.append(None)
        elif month <= nodes[0][0]:
            out.append(nodes[0][1])
        elif month >= nodes[-1][0]:
            out.append(nodes[-1][1])
        else:
            k = max(i for i in range(len(nodes)) if nodes[i][0] <= month)
            (m0, v0), (m1, v1) = nodes[k], nodes[k + 1]
            out.append(v0 + (v1 - v0) * (month - m0) / (m1 - m0))
    return out


def compute(panel, fsb, sofr, approved, history, band_bp=DEFAULT_BAND_BP,
            lookback=DEFAULT_LOOKBACK):
    """Returns a dict of per-term lists mirroring the dashboard's compute()."""
    avg = [_mean([offers[i] for _, offers in panel]) for i in range(NT)]
    n_quotes = [sum(offers[i] is not None for _, offers in panel) for i in range(NT)]
    floored = [max(a, f) if (a is not None and f is not None) else (a if a is not None else f)
               for a, f in zip(avg, fsb)]
    floor_binds = [a is not None and f is not None and f > a for a, f in zip(avg, fsb)]
    mtp = [round((fl - s) * 100, 1) if (fl is not None and s is not None) else None
           for fl, s in zip(floored, sofr)]
    lb = max(2, int(lookback))
    hist_inc = [vals for _, vals in history[-(lb - 1):]]
    trail_inc = []
    for i in range(NT):
        xs = [v[i] for v in hist_inc if v[i] is not None]
        if mtp[i] is not None:
            xs.append(mtp[i])
        trail_inc.append(round(mean(xs), 1) if xs else None)
    d_app = [round(m - a, 1) if m is not None else None for m, a in zip(mtp, approved)]
    d_trend = [round(t - a, 1) if t is not None else None for t, a in zip(trail_inc, approved)]
    status = []
    for i in range(NT):
        if mtp[i] is None:
            status.append("n/a")
        elif d_trend[i] is not None and abs(d_trend[i]) > band_bp:
            status.append("Discuss")
        elif abs(d_app[i]) > band_bp:
            status.append("Watch")
        else:
            status.append("Hold")
    return {"avg": avg, "n_quotes": n_quotes, "floored": floored,
            "floor_binds": floor_binds, "mtp": mtp, "trail_inc": trail_inc,
            "d_app": d_app, "d_trend": d_trend, "status": status,
            "band_bp": band_bp, "lookback": lb}

# ---------------------------------------------------------------- output ----

def write_results(out_path, C, sofr, fhlb, fsb, approved, as_of, dates):
    wb = Workbook()
    ws = wb.active
    ws.title = "Node summary"
    ws.append(["Term", "Market TP (bp)", f"{C['lookback']}w trend incl wk (bp)",
               "Approved TP (bp)", "Trend vs production (bp)", "Latest vs production (bp)",
               "Quotes", "Floor binds", "Status",
               "Panel avg (%)", "FSB print (%)", "Floored (%)", "SOFR (%)", "FHLB (%)"])
    for i in range(NT):
        ws.append([TERM_LABELS[i], C["mtp"][i], C["trail_inc"][i], approved[i],
                   C["d_trend"][i], C["d_app"][i], C["n_quotes"][i],
                   "yes" if C["floor_binds"][i] else "", C["status"][i],
                   C["avg"][i], fsb[i], C["floored"][i], sofr[i], fhlb[i]])
    ws2 = wb.create_sheet("Monthly curve")
    ws2.append(["Month", "Tenor", "SOFR (%)", "FHLB (%)", "Panel avg (%)",
                "Floored CD (%)", "Market TP (bp)", "Approved TP (bp)", "Production FTP (%)"])
    m_sofr, m_fhlb = interp_monthly(sofr), interp_monthly(fhlb)
    m_avg, m_floor = interp_monthly(C["avg"]), interp_monthly(C["floored"])
    m_app = interp_monthly(approved)
    for m in range(1, MMAX + 1):
        k = m - 1
        tp = round((m_floor[k] - m_sofr[k]) * 100, 1) if (m_floor[k] is not None and m_sofr[k] is not None) else None
        prod = round(m_sofr[k] + m_app[k] / 100, 5) if (m_sofr[k] is not None and m_app[k] is not None) else None
        label = f"{m // 12}Y" if m % 12 == 0 else f"{m}M"
        ws2.append([m, label, m_sofr[k], m_fhlb[k], m_avg[k], m_floor[k], tp, m_app[k], prod])
    ws3 = wb.create_sheet("Meta")
    for k, v in [("Week ending", as_of), ("Zone of indifference (bp)", C["band_bp"]),
                 ("Lookback (weeks)", C["lookback"]), ("CD panel date", dates.get("cd")),
                 ("FSB print date", dates.get("fsb")), ("SOFR date", dates.get("sofr")),
                 ("FHLB date", dates.get("fhlb"))]:
        ws3.append([k, str(v) if v else ""])
    wb.save(out_path)
