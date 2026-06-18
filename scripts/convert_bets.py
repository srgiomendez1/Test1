#!/usr/bin/env python3
"""Convert the friends' quiniela Excel into data/bets.json.

The workbook has one master "concentrado" sheet plus one tab per player. Each
player tab lists every group-stage match with that player's predicted score in
the "Goles Local" / "Goles Visitante" columns. We read those, map the Spanish
team names to openfootball English names, and key each match as
"YYYY-MM-DD|Team1|Team2" so predictions line up with the results feed.

Usage:
    python3 scripts/convert_bets.py <path-to-xlsx> [-o data/bets.json]
"""
import argparse
import json
import os
import sys
import urllib.request

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from team_aliases import to_en  # noqa: E402

OPENFOOTBALL_URL = (
    "https://raw.githubusercontent.com/openfootball/worldcup.json"
    "/master/2026/worldcup.json"
)


def load_fixtures():
    """Canonical fixtures from openfootball, indexed by the unordered team pair.

    The Excel and openfootball sometimes disagree on which team is "home", so we
    use openfootball's orientation as canonical and flip predictions to match.
    Group-stage pairs are unique, so a frozenset of the two names is a safe key.
    """
    req = urllib.request.Request(OPENFOOTBALL_URL, headers={"User-Agent": "wc-bet/1.0"})
    with urllib.request.urlopen(req, timeout=20) as r:
        feed = json.loads(r.read().decode("utf-8"))
    index = {}
    for m in feed.get("matches", []):
        h, a = m.get("team1"), m.get("team2")
        if not (h and a) or any(c.isdigit() for c in h + a):
            continue  # skip knockout placeholders (W101, 1A, ...)
        index[frozenset((h, a))] = {"date": m["date"], "home": h, "away": a}
    return index

# The master/summary sheet is not a player; everything else is.
MASTER_SHEET = "Quiniela Mundial 2026 - Grupos"

# Column indexes (0-based) within each player tab, taken from the header row.
COL_FECHA = 2
COL_LOCAL = 6
COL_GOLES_LOCAL = 7
COL_GOLES_VISITANTE = 8
COL_VISITANTE = 9

MONTHS = {
    # English abbreviations
    "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05",
    "jun": "06", "jul": "07", "aug": "08", "sep": "09", "oct": "10",
    "nov": "11", "dec": "12",
    # Spanish abbreviations (just in case)
    "ene": "01", "abr": "04", "ago": "08", "dic": "12",
}


def parse_date(raw):
    """'11-Jun-2026' -> '2026-06-11'. Also accepts datetime/date objects."""
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    parts = s.replace("/", "-").split("-")
    if len(parts) != 3:
        return None
    day, mon, year = parts
    mon = MONTHS.get(mon.strip().lower()[:3], mon)
    return f"{year}-{int(mon):02d}-{int(day):02d}"


def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def convert(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    players = [s for s in wb.sheetnames if s.strip() != MASTER_SHEET]
    fixtures = load_fixtures()

    predictions = {}      # match_key -> {player: [h, a]}
    match_meta = {}       # match_key -> {date, home, away} (English names)
    unmapped = set()
    unmatched = set()

    for player in players:
        ws = wb[player]
        name = player.strip()
        for row in ws.iter_rows(values_only=True):
            local = row[COL_LOCAL]
            visit = row[COL_VISITANTE]
            if not (local and visit) or parse_date(row[COL_FECHA]) is None:
                continue
            try:
                en_local = to_en(str(local).strip())
                en_visit = to_en(str(visit).strip())
            except KeyError as e:
                unmapped.add(str(e))
                continue

            # Align to openfootball's canonical orientation (home/away may differ).
            fx = fixtures.get(frozenset((en_local, en_visit)))
            if not fx:
                unmatched.add(f"{en_local} vs {en_visit}")
                continue
            key = f"{fx['date']}|{fx['home']}|{fx['away']}"
            match_meta.setdefault(key, dict(fx))

            gl = to_int(row[COL_GOLES_LOCAL])
            gv = to_int(row[COL_GOLES_VISITANTE])
            if gl is None or gv is None:
                continue  # player left this match blank
            # If the Excel's home team is openfootball's away team, flip the score.
            pred = [gl, gv] if en_local == fx["home"] else [gv, gl]
            predictions.setdefault(key, {})[name] = pred

    if unmatched:
        print("WARNING: Excel matches with no openfootball fixture:", file=sys.stderr)
        for u in sorted(unmatched):
            print("  -", u, file=sys.stderr)
    if unmapped:
        print("ERROR: unmapped team names found:", file=sys.stderr)
        for u in sorted(unmapped):
            print("  -", u, file=sys.stderr)
        sys.exit(1)

    return {
        "players": [p.strip() for p in players],
        "matches": match_meta,
        "predictions": predictions,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("-o", "--out", default="data/bets.json")
    args = ap.parse_args()

    data = convert(args.xlsx)
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    n_preds = sum(len(v) for v in data["predictions"].values())
    print(f"Wrote {args.out}: {len(data['players'])} players, "
          f"{len(data['matches'])} matches, {n_preds} predictions.")


if __name__ == "__main__":
    main()
